import sys
from clingo.application import Application, clingo_main
from clingo import Number
import pprint

import calendar
import textwrap as _textwrap
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill, NamedStyle
import itertools

import os

from enum import Enum

DAY = "day"
NIGHT = "night"
WEEKDAY = "weekday"
SAT = "sat"
SUN = "sun"

DATA_FOLDER = "data"
PERSONS_FILE = DATA_FOLDER + os.sep + "nombres"
SEPARATED_FILE = DATA_FOLDER + os.sep +"no-juntos"
DAYS_BLOCKED_FILE = DATA_FOLDER + os.sep +"dias-bloqueados"
CONSTRAINTS_FILE = DATA_FOLDER + os.sep +"especiales"

# styles
Text_14_bold = Font(name='Calibri', size=14, bold=True)
Text_12_bold = Font(name='Calibri', size=12, bold=True)
Text_12 = Font(name='Calibri', size=12)
Text_12_bold_red = Font(name='Calibri', color="00FF0000", size=12, bold=True)

center_aligned_text = Alignment(horizontal="center")

yellow_fill = PatternFill(fill_type="solid", start_color="00FFFF00", end_color="00FFFF00")
orange_fill = PatternFill(fill_type="solid", start_color="00FA9523", end_color="00FA9523")
green_fill = PatternFill(fill_type="solid", start_color="008FF54C", end_color="008FF54C")
brown_fill = PatternFill(fill_type="solid", start_color="00FFCC99", end_color="00FFCC99")

border_top_bottom = Border(top=Side(style="medium"), bottom=Side(style="medium"))
border_top_bottom_left = Border(top=Side(style="medium"), bottom=Side(style="medium"), left=Side(style="medium"))
border_top_bottom_right = Border(top=Side(style="medium"), bottom=Side(style="medium"), right=Side(style="medium"))

# full styles
weekday_style = NamedStyle(name=WEEKDAY)
weekday_style.font = Text_14_bold
weekday_style.alignment = center_aligned_text

normal_bold_style = NamedStyle(name="normal_bold")
normal_bold_style.font = Text_12_bold
normal_bold_style.alignment = center_aligned_text

normal_style = NamedStyle(name="normal")
normal_style.font = Text_12
normal_style.alignment = center_aligned_text

normal_red_style = NamedStyle(name="normal_red")
normal_red_style.font = Text_12_bold_red
normal_red_style.alignment = center_aligned_text


def xlref(row, column, zero_indexed=False):
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)

class ClingoApp(Application):
	def __init__(self, name):
		self.version = "0.1.5"
		self.program_name = name
		self.pp = pprint.PrettyPrinter(indent=4)

		self.month = None
		self.year = None
		self.holidays = []

	def __on_model(self, model):
		self.model = model
		cal = self.parse_model()
		self.create_workbook(cal)

	def parse_model(self):
		cal = {}
		self.total_days = {}
		self.total_shifts = {}
		self.hours_per_person_per_shift = {}

		for atom in self.model.symbols(atoms=True):
			if atom.name == "assigned":

				person = atom.arguments[1].name
				person = person.capitalize()
				day_num = atom.arguments[2].number

				shift = atom.arguments[0].name

				cal.setdefault(day_num, {DAY: [], NIGHT: [], "type,week": (self.day_to_weekday[day_num] ,self.day_to_week[day_num])})[shift].append(person)

				self.total_days.setdefault(person, 0)
				self.total_days[person] = self.total_days[person] + 1

				self.total_shifts.setdefault(person, {WEEKDAY: {DAY: 0, NIGHT: 0},
								SAT: {DAY: 0, NIGHT: 0},
								SUN: {DAY: 0, NIGHT: 0}	})


				self.total_shifts[person][self.get_day_type(day_num)][shift] += 1


				self.hours_per_person_per_shift.setdefault(person, []).append(self.hours_per_day[day_num][shift])

		return cal

	def create_workbook(self, cal):
		workbook = Workbook()

		self.create_main_sheet(cal, workbook.active)

		#sheet_hours = workbook.create_sheet("hours")
		self.create_hours_table(cal, workbook.active, r_offset=self.get_week_count()*5+5, c_offset=2)
		self.create_days_per_person_table(workbook.active, r_offset=2, c_offset=12)
		self.create_hours_per_shift_table(workbook.active, r_offset=self.get_week_count()*5, c_offset=12)

		self.set_sheet_dims(workbook.active)

		workbook.save(filename=f"{self.month}-{self.year}.xlsx")

	def set_sheet_dims(self, sheet):
		dims = {}
		for row in sheet.rows:
			for cell in row:
				if cell.value:
					dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
		for col, value in dims.items():
			sheet.column_dimensions[col].width = value + 3

	def is_weekday_row(self, row_offset, row):

		return (row -  row_offset) % 5 == 0

	def create_main_sheet(self, cal, sheet):
		row_offset = 4
		col_offset = 4

		shift_offset = {DAY: 1, NIGHT: 3}

		cell = sheet.cell(row=row_offset-2, column=col_offset-1)
		cell.value = f"Turno para el mes {self.month} del año {self.year}"
		cell.font = Text_14_bold
		cell.alignment = center_aligned_text

		cell = sheet.cell(row=row_offset-1, column=col_offset-1)
		cell.value = "Horario"
		cell.style = weekday_style

		cell = sheet.cell(row=row_offset-1, column=col_offset-2)
		cell.value = "Turno"
		cell.style = weekday_style

		weekdaynum_to_weekday = {0: "Lunes", 1: "Martes", 2: "Miercoles", 3: "Jueves", 4: "Viernes", 5: "Sabado", 6: "Domingo"}
		for i in range(0,7):
			cell = sheet.cell(row=row_offset-1, column=col_offset+i)
			cell.value = weekdaynum_to_weekday[i]
			cell.style = weekday_style

		for day in sorted(cal.keys()):
			weekday,week = cal[day]["type,week"]

			cell = sheet.cell(row=row_offset + week*5, column=col_offset+weekday)
			cell.value = day
			cell.style = normal_bold_style
			if weekday >= 5:
				cell.font = Text_12_bold_red
				
			for shift in [DAY, NIGHT]:
				for order, person in enumerate(cal[day][shift], start=0):
					row = row_offset + self.day_to_week[day]*5 + shift_offset[shift] + order
					col = col_offset + self.day_to_weekday[day]
					cell = sheet.cell(row=row, column=col)
					cell.value = person
					cell.style = normal_bold_style
					if shift == DAY:
						cell.fill = brown_fill

					if day in self.holidays:
						cell.font = Text_12_bold_red
					if weekday >= 5:
						cell.font = Text_12_bold_red
						cell.fill = yellow_fill


					cell = sheet.cell(row=row, column=col_offset-1)
					cell.value = "7:00pm - 7:00am"
					cell.style = normal_bold_style
					if shift == DAY:
						cell.value = "7:00am - 7:00pm"
						cell.fill = brown_fill

					cell = sheet.cell(row=row, column=col_offset-2)
					cell.value = shift
					cell.style = normal_bold_style
					if shift == DAY:
						cell.fill = brown_fill

		# sum of shifts in the week

		# add row offset -2 and -1 so it has the headers
		# the rest is the weekday numbers
		for row in range(row_offset+1, row_offset + (self.get_week_count())*5):
			if self.is_weekday_row(row_offset, row):
				continue
			cell = sheet.cell(row=row, column=col_offset+7)
			first = xlref(row, col_offset)
			last = xlref(row, col_offset+6)
			cell.value = f"=COUNTA({first}:{last})"
			cell.fill = orange_fill
		
		cell = sheet.cell(row=row_offset + (self.get_week_count()*5) + 1, column=col_offset+7)
		first = xlref(row_offset+1, col_offset+7)
		last = xlref(row_offset + (self.get_week_count()*5) - 1, col_offset+7)
		cell.value = f"=SUM({first}:{last})"
		cell.fill = green_fill
	
		# border of cells
		# *5 since every "row" has the weekday + 2 day shifts + 2 night shifts = 5 actual rows
		for row in range(row_offset-2, row_offset + (self.get_week_count())*5):
			for col in range(col_offset-2, col_offset+7):
				cell = sheet.cell(row=row, column=col)

				if self.is_weekday_row(row_offset, row):
					if col == col_offset-2:
						cell.border = border_top_bottom_left
					elif col == col_offset+6:
						cell.border = border_top_bottom_right
					cell.border = border_top_bottom
				
				elif not self.is_weekday_row(row_offset, row):
					if col == col_offset-2:
						cell.border = Border(left=Side("medium"))
					elif col == col_offset+6:
						cell.border = Border(right=Side("medium"))

	def create_hours_table(self, cal, sheet, r_offset=0, c_offset=0):
		row_offset = 4 + r_offset
		col_offset = 2 + c_offset

		shift_offset = {DAY: 1, NIGHT: 3}


		weekdaynum_to_weekday = {0: "Lunes", 1: "Martes", 2: "Miercoles", 3: "Jueves", 4: "Viernes", 5: "Sabado", 6: "Domingo"}
		for i in range(0,7):
			cell = sheet.cell(row=row_offset-1, column=col_offset+i)
			cell.value = weekdaynum_to_weekday[i]
			cell.style = weekday_style


		for day in sorted(cal.keys()):
			weekday,week = cal[day]["type,week"]

			cell = sheet.cell(row=row_offset + week*5, column=col_offset+weekday)
			cell.value = day
			cell.style = normal_bold_style
			cell.fill = yellow_fill
				
			for shift in [DAY, NIGHT]:
				for order, person in enumerate(cal[day][shift], start=0):
					row = row_offset + self.day_to_week[day]*5 + shift_offset[shift] + order
					col = col_offset + self.day_to_weekday[day]
					cell = sheet.cell(row=row, column=col)
					cell.value = self.hours_per_day[day][shift]
					cell.style = normal_style
					if shift == DAY:
						cell.fill = brown_fill

					if day in self.holidays:
						cell.font = Text_12_bold_red

		for row in range(row_offset+1, row_offset + (self.get_week_count())*5):
			if self.is_weekday_row(row_offset, row):
				continue
			cell = sheet.cell(row=row, column=col_offset+7)
			first = xlref(row, col_offset)
			last = xlref(row, col_offset+6)
			cell.value = f"=SUM({first}:{last})"
			cell.fill = orange_fill
		
		cell = sheet.cell(row=row_offset + (self.get_week_count()*5) + 1, column=col_offset+7)
		first = xlref(row_offset+1, col_offset+7)
		last = xlref(row_offset + (self.get_week_count()*5) - 1, col_offset+7)
		cell.value = f"=SUM({first}:{last})"
		cell.fill = green_fill

	def create_days_per_person_table(self, sheet, r_offset=0, c_offset=0):
		row_offset = 2 + r_offset
		col_offset = 3 + c_offset

		# header
		cell = sheet.cell(row=row_offset, column=col_offset)
		cell.value = "APELLIDO"
		cell.fill = orange_fill
		cell.font = Text_12

		for col, daytype_shift in enumerate(itertools.product(["SE", "S", "D"], ["D", "N"]), start=1):
			day_type, shift = daytype_shift
			cell = sheet.cell(row=row_offset, column=col + col_offset)
			cell.value = f"{day_type}-{shift}"
			cell.fill = orange_fill
			cell.font = Text_12

		cell = sheet.cell(row=row_offset, column=col + col_offset + 1)
		cell.value = "SUMA"	
		cell.fill = orange_fill
		cell.font = Text_12

		# actual values
		for row, person in enumerate(self.total_shifts.keys(), start=1):
			cell = sheet.cell(row=row+row_offset, column=col_offset)
			cell.value = person
			cell.fill = green_fill
			cell.font = Text_12

			for col, daytype_shift in enumerate(itertools.product([WEEKDAY, SAT, SUN], [DAY, NIGHT]), start=1):
					day_type, shift = daytype_shift
					cell = sheet.cell(row=row+row_offset, column=col + col_offset)
					cell.value = self.total_shifts[person][day_type][shift]
					cell.fill = green_fill
					cell.font = Text_12

			# cell for the sum of the total days of the person
			# we make use of the variable from the iterator(col) to know the amount of columns we have
			cell = sheet.cell(row=row+row_offset, column=col + col_offset + 1)
			first = xlref(row+row_offset, col_offset + 1)
			last =  xlref(row+row_offset, col_offset + col)
			cell.value = f"=SUM({first}:{last})"
			cell.fill = green_fill
			cell.font = Text_12
		
		# cell of the sum of total days
		# we make use of the iterator variables (col and row) to know the max row and cols
		cell = sheet.cell(row=row+row_offset+1, column=col + col_offset + 1)
		first = xlref(row_offset+1, col_offset + col + 1 )
		last =  xlref(row+row_offset, col_offset + col + 1)
		cell.value = f"=SUM({first}:{last})"
		cell.fill = brown_fill
		cell.font = Text_12

	def create_hours_per_shift_table(self, sheet, r_offset=0, c_offset=0):
		row_offset = 2 + r_offset
		col_offset = 3 + c_offset

		MAX_COLS = 13

		cell = sheet.cell(row=row_offset-1, column=col_offset)
		cell.value = "APELLIDO"
		cell.fill = orange_fill
		cell.font = Text_12

		for col in range(1,MAX_COLS):
			cell = sheet.cell(row=row_offset-1, column=col_offset+col)
			cell.value = col
			cell.fill = orange_fill
			cell.font = Text_12

		cell = sheet.cell(row=row_offset-1, column=col_offset+col+1)
		cell.value = "SUMA"
		cell.fill = orange_fill
		cell.font = Text_12

		for row, person in enumerate(self.hours_per_person_per_shift.keys(), start=0):
			# name cell
			cell = sheet.cell(row=row_offset+row, column=col_offset)
			cell.value = person
			if row % 2:
				cell.font = Text_12_bold_red
			else:
				cell.font = Text_12_bold

			# value cells
			for col, hours in enumerate(self.hours_per_person_per_shift[person], start=1):
				cell = sheet.cell(row=row_offset+row, column=col_offset+col)
				cell.value = hours
				if row % 2:
					cell.font = Text_12_bold_red
				else:
					cell.font = Text_12_bold

			# sum cell
			cell = sheet.cell(row=row_offset+row, column=col_offset+MAX_COLS)
			first = xlref(row_offset+row, col_offset+1)
			last = xlref(row_offset+row, col_offset+col)
			cell.value = f"=SUM({first}:{last})"
			if row % 2:
				cell.font = Text_12_bold_red
			else:
				cell.font = Text_12_bold


		# total sum cell 
		cell = sheet.cell(row=row_offset+row+1, column=col_offset+MAX_COLS)
		first = xlref(row_offset, col_offset+MAX_COLS)
		last = xlref(row_offset+row, col_offset+MAX_COLS)
		cell.value = f"=SUM({first}:{last})"
		cell.font = Text_12_bold
		cell.fill = brown_fill

	def register_options(self, options):
		"""
		See clingo.clingo_main().
		"""

		group = "Turno Options"

		options.add(group, "month", _textwrap.dedent("""Month as a number"""), self.__parse_month)
		options.add(group, "year", _textwrap.dedent("""Year"""), self.__parse_year)
		options.add(group, "holiday", _textwrap.dedent("""Holidays separated by ,"""), self.__parse_holiday)

	def __parse_holiday(self, option):
		self.holidays = [int(h) for h in option.split(",")]
		
		for h in self.holidays:
			if h > 31 or h < 1:
				return False
		
		return True

	def __parse_month(self, option):
		self.month = int(option)
		
		if self.month < 13 and self.month > 0:
			return True
		
		print("Month has to be in the range 1 to 12")
		
		return False

	def __parse_year(self, option):
		self.year = int(option)
		
		if self.year < 2100 and self.year > 2000:
			return True
		
		print("year has to be in the range 2000 to 2100")
		
		return False

	def check_not_none(self, val, msg):
		if val is None:
			print(msg)
			return False
		return True

	def min_max(self, values):

		days, fixed_days, persons, fixed_persons = values.arguments

		val = ((days.number * 4) - fixed_days.number) / (persons.number - fixed_persons.number)
		return Number(int(val))

	def load_info_files(self, ctl):
		persons = self.parse_persons()
		blocks = self.parse_blocks()
		separated = self.parse_separated()
		constraints = self.parse_constraints()

		print(persons)
		print(blocks)
		print(separated)

		ctl.add("base", [], persons)
		ctl.add("base", [], blocks)
		ctl.add("base", [], separated)
		ctl.add("base", [], constraints)
		
	def parse_persons(self):
		persons = ""
		with open(PERSONS_FILE, "r") as _f:
			for person in _f.readlines():
				persons += f"person({person.strip()}).\n"

		return persons

	def parse_blocks(self):
		blocks = ""
		with open(DAYS_BLOCKED_FILE, "r") as _f:
			for block in _f.readlines():
				person, first, last = str(block.strip()).split(",")
				blocks += f"constraint(blocked, ({person}, {first}..{last})).\n"

		return blocks

	def parse_separated(self):
		separated = ""
		with open(SEPARATED_FILE, "r") as _f:
			for person in _f.readlines():
				separated += f"single({person.strip()}).\n"

		return separated

	def parse_constraints(self):
		constraints = ""
		with open(CONSTRAINTS_FILE, "r") as _f:
			for constraint in _f.readlines():
				constraint = constraint.strip().split(",")
				if len(constraint) == 0:
					continue
				if constraint[0] == "constraint":
					constraints += f"constraint({constraint[1]}, ({','.join(constraint[2:])})).\n"
				
				elif constraint[0] == "exception":
					constraints += f"exception({constraint[1]}, ({','.join(constraint[2:])})).\n"

				elif constraint[0] == "special_days":
					constraints += f"special_days({constraint[1]},{constraint[2]}).\n"

		return constraints

	def get_week_count(self):
		return len(calendar.Calendar().monthdays2calendar(self.year, self.month))

	def get_day_type(self, day):
		if self.day_to_weekday[day] < 5 and self.day_to_weekday[day] >= 0:
			return WEEKDAY
		elif self.day_to_weekday[day] == 5:
			return SAT
		elif self.day_to_weekday[day] == 6:
			return SUN
		else:
			raise ValueError(f"Value {day} is an invalid number for a weekday")

	def main(self, ctl, files):

		if self.month is None:
			month = input("Mes: ")
			try:
				self.month = int(month)
			except:
				raise ValueError(f"{month} no es un mes valido")

			if self.month > 12 or self.month < 1:
				raise ValueError(f"{month} no es un mes valido")

		if self.year is None:
			year = input("Año: ")
			try:
				self.year = int(year)
			except:
				raise ValueError(f"{year} no es un año valido")
			

		if len(self.holidays) == 0:
			holidays = input("Dias Feriados: ")
			if holidays != "":
				try:
					self.holidays = [int(day) for day in holidays.split(",")]
				except:
					raise ValueError(f"{day} no es un dia valido")

				for day in self.holidays:
					if day < 1 or day > 31:
						raise ValueError(f"{day} no es un dia valido!")
			else:
				holidays = []

		ctl.load("turno.lp")
		for f in files:
			if "turno.lp" not in f:
				ctl.load(f)

		self.hours_per_day = {}

		self.day_to_weekday = {}
		self.day_to_week = {}

		self.days_in_week = {}

		dates = calendar.monthrange(self.year, self.month)[1]
		for i, week in enumerate(calendar.Calendar().monthdays2calendar(self.year, self.month)):
			self.days_in_week[i] = len(week)
			# day is the date, weekday is in numbers if its monday, tuesday etc. 5 is saturday, 6 in sunday
			for day, weekday in week:
				if day == 0:
					continue
				self.day_to_weekday[day] = weekday
				self.day_to_week[day] = i

				if day in self.holidays:
					day_hours = 12
					night_hours = 12
				elif self.get_day_type(day) == SAT:
					day_hours = 9
					night_hours = 9
				elif  self.get_day_type(day) == SUN:
					day_hours = 12
					night_hours = 12
				else:
					day_hours = 3
					night_hours = 4

				self.hours_per_day[day] = {DAY: day_hours,
									  NIGHT: night_hours}

		atoms = ""
		max_hours = 0
		for day, shifts in self.hours_per_day.items():
			for shift, hours in shifts.items():
				atoms += f"hours({shift},{day},{hours}).\n"

				max_hours += hours*2
		
		self.load_info_files(ctl)

		ctl.add("base", [], atoms)

		ctl.add("base", [], f"max_hours({max_hours}).\n")

		s1 = ".\n".join([f"day_to_weekday({day},{weekday})" for day, weekday in self.day_to_weekday.items()]) + ".\n"
		ctl.add("base", [], s1)
		
		s2 = ".\n".join([f"day_to_week({day},{week})" for day, week in self.day_to_week.items()]) + ".\n"
		ctl.add("base", [], s2)

		ctl.add("base", [], f"max_days({dates}).")

		ctl.ground([("base", [])], self)
		ctl.solve(on_model=self.__on_model)

		cal = self.parse_model()

clingo_main(ClingoApp(sys.argv[0]), sys.argv[1:])